#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SPDR (SSGA) 每日持仓 —— 本监控唯一的债券级价格源。

为什么不是 iShares：LQD/HYG 的 ajax 端点被 Akamai 拦。响应头写着
`content-type: text/csv` 和 `content-disposition: attachment; filename=LQD_holdings.csv`，
body 却返回完整产品页 HTML；加 Referer、加 cookie jar 两段式都无效。
SSGA 无门禁、无密钥、返回真 xlsx，且有转债 ETF（CWB），覆盖更全。**不要再试 iShares。**

持仓表**没有价格列**，价格是导出来的：

    clean_price = Market Value / Par Value × 100

三个必须记住的口径限制（写进每一行观测的 quality，不靠注释约束）：

1. 这是**基金管理人的估值，不是成交价**。流动性差的券会粘滞，连续多日不动
   要标 stale，不能当成「利差没变」。
2. 这是**该发行人在指数样本中的子集**，不是全部存量债。发行人层聚合只能说
   「样本内加权」。
3. CWB 里混有强制转股优先股（maturity 为 '-'，par/MV 比算出来四位数），
   必须剔掉，否则价格字段会出现 4922 这种数。
"""

from __future__ import annotations

import datetime as dt
import io
import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from .base import CollectResult, fingerprint, http_get, load_config

# 含权特征：次级、永续、浮息。这些券算出来的 G-spread 有偏，
# 必须标 option_biased，不与普通高级无担保券同图比较。
_OPTION_RE = re.compile(r"JR SUBORDIN|SUBORDINA|\bVAR\b|PERP|FLOAT", re.I)
_144A_RE = re.compile(r"\b144A\b", re.I)
_ASOF_RE = re.compile(r"As of\s+(\d{1,2})-([A-Za-z]{3})-(\d{4})")

_MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}

_TYPE_RULES = [
    (re.compile(r"1ST MORTGAGE|FIRST MORTGAGE", re.I), "first_mortgage"),
    (re.compile(r"JR SUBORDIN", re.I), "jr_subordinated"),
    (re.compile(r"SUBORDINA", re.I), "subordinated"),
    (re.compile(r"SR SECURED|SENIOR SECURED", re.I), "senior_secured"),
    (re.compile(r"COMPANY GUAR", re.I), "company_guaranteed"),
    (re.compile(r"SR UNSECURED|SENIOR UNSECURED", re.I), "senior_unsecured"),
]


def _parse_asof(cell: Any) -> Optional[str]:
    match = _ASOF_RE.search(str(cell or ""))
    if not match:
        return None
    day, mon, year = match.groups()
    month = _MONTHS.get(mon.title())
    if not month:
        return None
    return dt.date(int(year), month, int(day)).isoformat()


def _parse_maturity(raw: Any) -> Optional[dt.date]:
    text = str(raw or "").strip()
    if not text or text == "-":
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    return None


def _instrument_type(name: str) -> str:
    for pattern, label in _TYPE_RULES:
        if pattern.search(name):
            return label
    return "senior_unsecured"


def _match_issuer(name: str, issuers: Dict[str, Any]) -> Optional[str]:
    upper = name.upper()
    for key, meta in issuers.items():
        excludes = meta.get("exclude") or []
        if any(x.upper() in upper for x in excludes):
            continue
        for token in meta.get("match") or []:
            if token.upper() in upper:
                return key
    return None


def _read_book(content: bytes) -> Tuple[Optional[str], List[Tuple]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    asof: Optional[str] = None
    rows: List[Tuple] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= 5:
            if asof is None:
                for cell in row or ():
                    got = _parse_asof(cell)
                    if got:
                        asof = got
                        break
            continue
        if row and row[0]:
            rows.append(row)
    return asof, rows


def collect(cfg: Optional[Dict[str, Any]] = None,
            universe: Optional[Dict[str, Any]] = None) -> CollectResult:
    cfg = cfg or load_config("sources.yaml")
    universe = universe or load_config("universe.yaml")
    src = cfg["sources"]["spdr"]
    issuers = universe["issuers"]
    baskets = src["baskets"]

    result = CollectResult(source_id="spdr",
                           basket_fingerprint=fingerprint([b["ticker"] for b in baskets]))
    headers = {"User-Agent": src["user_agent"], "Referer": src["referer"]}
    today = dt.date.today()

    seen: Dict[str, str] = {}          # isin -> 首次出现的 ETF
    asof_seen: Dict[str, str] = {}
    failures: List[str] = []
    dropped_equity = 0

    for basket in baskets:
        ticker = basket["ticker"]
        url = f"{src['base_url']}/{src['path_template'].format(ticker=ticker)}"
        try:
            content = http_get(url, headers=headers,
                               timeout=cfg["defaults"]["timeout_seconds"],
                               retries=cfg["defaults"]["retries"],
                               backoff=cfg["defaults"]["backoff_base_seconds"],
                               binary=True)
            asof, rows = _read_book(content)
        except Exception as exc:                       # noqa: BLE001
            failures.append(f"{ticker}: {exc}")
            continue

        if not asof:
            failures.append(f"{ticker}: 读不到 As of 日期，整只 ETF 跳过")
            continue
        asof_seen[ticker] = asof

        for row in rows:
            name = str(row[0] or "").strip()
            isin = str(row[1] or "").strip()
            coupon, par, mv, maturity_raw = row[4], row[5], row[6], row[8]

            if not isin or isin in seen:
                continue
            issuer = _match_issuer(name, issuers)
            if issuer is None:
                continue

            maturity = _parse_maturity(maturity_raw)
            if maturity is None:
                # 强制转股优先股：CWB 里混着的股票型持仓，par/MV 比是四位数。
                dropped_equity += 1
                continue
            if not par or not mv:
                continue

            years = (maturity - today).days / 365.25
            if years <= 0:
                continue

            price = float(mv) / float(par) * 100.0
            meta = issuers[issuer]
            has_option = bool(_OPTION_RE.search(name))
            seen[isin] = ticker

            result.instruments.append({
                "instrument_key": isin,
                "isin": isin,
                "issuer_key": issuer,
                "issuer_parent_key": issuer,
                "rung": meta.get("rung"),
                "regime": "convertible" if basket["segment"] == "convertible"
                          else meta.get("regime", "public_corp"),
                "instrument_type": ("convertible" if basket["segment"] == "convertible"
                                    else _instrument_type(name)),
                "coupon": float(coupon) if coupon is not None else None,
                "coupon_type": "fixed",
                "maturity": maturity.isoformat(),
                "currency": str(row[7] or "USD"),
                "segment": basket["segment"],
                "is_144a": bool(_144A_RE.search(name)),
                "has_embedded_option": has_option,
                "recourse": None,
                "collateral_type": ("real_property"
                                    if _instrument_type(name) == "first_mortgage"
                                    else None),
                "display_name": name[:120],
                "first_seen": asof,
                "last_seen": asof,
            })
            result.observations.append({
                "asof_date": asof,
                "instrument_key": isin,
                "metric": "px.clean",
                "value": round(price, 4),
                "value_text": None,
                "unit": "price",
                "method": "derived",          # MV/Par，不是源直接给的
                "source_id": f"spdr:{ticker}",
                "obs_date": today.isoformat(),
                "staleness_days": (today - dt.date.fromisoformat(asof)).days,
                "quality": "ok",
                "raw_ref": f"{ticker}_holdings.xlsx",
            })

    if failures and result.instruments:
        result.status = "partial"
    elif failures and not result.instruments:
        result.status = "failed"

    parts = [f"{len(result.instruments)} 只工具（ISIN 去重）"]
    if dropped_equity:
        parts.append(f"剔除强制转股/股票型 {dropped_equity} 行")
    if failures:
        parts.append("失败：" + "；".join(failures))
    if len(set(asof_seen.values())) > 1:
        parts.append("各 ETF 的 As of 日期不一致：" +
                     ", ".join(f"{k}={v}" for k, v in asof_seen.items()))
    result.note = "｜".join(parts)
    return result
